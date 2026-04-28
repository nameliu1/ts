import os
import re
import subprocess
import sys
import time
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter

HTTP_PORTS = {
    "80", "81", "82", "83", "85", "88", "8000", "8001", "8002", "8003",
    "8004", "8005", "8006", "8007", "8008", "8009", "8010", "8011", "8018",
    "8022", "8080", "8081", "8082", "8083", "8084", "8085", "8086", "8087",
    "8088", "8089", "8090", "8091", "8092", "8094", "8098", "8099", "8180",
    "8181", "8189", "8191", "8200", "8280", "888", "8888", "9000", "9001",
    "9002", "9003", "9004", "9005", "9006", "9007", "9010", "9011", "9012",
    "9100", "9200"
}

HTTPS_PORTS = {
    "443", "441", "444", "445", "1443", "3443", "4430", "4433", "4443", "4899",
    "5443", "7443", "8443", "8843", "9043", "9443", "9543", "10443", "11443",
    "12443", "13443", "14433"
}

def wait_for_file(file_path, timeout=60, min_size=10):
    """等待文件生成并达到最小大小"""
    import time
    start_time = time.time()
    while time.time() - start_time < timeout:
        if os.path.exists(file_path):
            size = os.path.getsize(file_path)
            if size > min_size:
                return True
        time.sleep(1)
    return False

def get_config_output_path():
    """从config.yaml读取配置的输出路径"""
    try:
        with open("config.yaml", "r", encoding="utf-8") as f:
            for line in f:
                if line.startswith("CfgOutPath:"):
                    path = line.split(":", 1)[1].strip()
                    if path and os.path.isdir(path):
                        return path
    except:
        pass
    return None

def run_ts_scan():
    """执行ts命令进行扫描，使用原生DOS界面显示输出"""
    import shutil
    print("开始执行端口扫描...")
    cmd = 'ts -hf ip.txt -portf ports.txt -np -m port,url'
    
    # 获取配置输出路径
    config_out_path = get_config_output_path()
    print(f"配置输出路径: {config_out_path}")
    print(f"当前工作目录: {os.getcwd()}")
    
    try:
        result = subprocess.run(cmd, shell=True, check=True, text=True)
        
        # 等待文件生成完成
        print("等待扫描结果写入文件...")
        time.sleep(3)  # 额外等待确保文件写入
        
        # 列出可能的url.txt位置
        possible_paths = [
            "url.txt",
            "port.txt", 
            os.path.join(config_out_path, "url.txt") if config_out_path else None,
            os.path.join(config_out_path, "port.txt") if config_out_path else None,
        ]
        
        print("=== 调试信息: 检查文件 ===")
        for p in possible_paths:
            if p and os.path.exists(p):
                size = os.path.getsize(p)
                print(f"  发现文件: {p}, 大小: {size} 字节")
            elif p:
                print(f"  不存在: {p}")
        
        # 检查并复制 url.txt
        url_found = False
        if os.path.exists("url.txt") and os.path.getsize("url.txt") > 10:
            url_found = True
            print(f"当前目录 url.txt 有效, 大小: {os.path.getsize('url.txt')}")
        elif config_out_path:
            config_url = os.path.join(config_out_path, "url.txt")
            if os.path.exists(config_url):
                shutil.copy2(config_url, "url.txt")
                print(f"从配置目录复制 url.txt: {config_url}")
                url_found = True
        
        if not url_found:
            print("警告: url.txt 未生成或为空，尝试从port.txt提取...")
            extract_urls_from_port_file()
        
        # 检查并复制 port.txt
        port_found = False
        if os.path.exists("port.txt") and os.path.getsize("port.txt") > 10:
            port_found = True
        elif config_out_path:
            config_port = os.path.join(config_out_path, "port.txt")
            if os.path.exists(config_port):
                shutil.copy2(config_port, "port.txt")
                print(f"从配置目录复制 port.txt: {config_port}")
                port_found = True
        
        if not port_found:
            print("警告: port.txt 未生成或为空")
        
        print("扫描完成，结果已保存到url.txt和port.txt")
        return True
    except subprocess.CalledProcessError as e:
        print(f"扫描失败，错误代码: {e.returncode}")
        return False
    except Exception as e:
        print(f"执行命令时发生错误: {e}")
        return False

def build_candidate_url(host, port, protocol_hint=""):
    """根据主机和端口推断 HTTP(S) URL。"""
    host = host.strip().strip(',')
    port = str(port).strip()
    protocol_hint = protocol_hint.upper()

    if not host or not port.isdigit():
        return None

    if "HTTPS" in protocol_hint or port in HTTPS_PORTS:
        scheme = "https"
    elif "HTTP" in protocol_hint or port in HTTP_PORTS:
        scheme = "http"
    else:
        return None

    if (scheme == "http" and port == "80") or (scheme == "https" and port == "443"):
        return f"{scheme}://{host}"

    return f"{scheme}://{host}:{port}"


def extract_urls_from_port_file():
    """从port.txt中提取URL信息，生成url.txt"""
    if not os.path.exists("port.txt"):
        return []

    urls = []
    with open("port.txt", "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            matches = re.findall(r'(https?://[^\s,\]\[]+)', line)
            for match in matches:
                url = match.rstrip(',').rstrip('/')
                if url and url not in urls:
                    urls.append(url)

            status_match = re.match(r'^([^:\s,\[\]]+):(\d+)\s+open$', line, re.IGNORECASE)
            if status_match:
                candidate = build_candidate_url(status_match.group(1), status_match.group(2))
                if candidate and candidate not in urls:
                    urls.append(candidate)
                continue

            protocol_match = re.match(
                r'^([A-Z/]+),.*?,\s*([^:\s,\[\]]+):(\d+),',
                line,
                re.IGNORECASE,
            )
            if protocol_match:
                candidate = build_candidate_url(
                    protocol_match.group(2),
                    protocol_match.group(3),
                    protocol_match.group(1),
                )
                if candidate and candidate not in urls:
                    urls.append(candidate)

    if urls:
        with open("url.txt", "w", encoding="utf-8") as f:
            for url in urls:
                f.write(url + "\n")
        print(f"从port.txt提取了 {len(urls)} 个URL到url.txt")

    return urls

def parse_url_file():
    """解析url.txt文件，提取详细信息，处理URL末尾的逗号"""
    if not os.path.exists("url.txt") or os.path.getsize("url.txt") < 10:
        print("url.txt文件不存在或为空，尝试从port.txt提取URL...")
        urls = extract_urls_from_port_file()
        if not urls:
            print("没有解析到任何URL数据！")
            return []
    
    if not os.path.exists("url.txt"):
        print("url.txt文件不存在，请先执行扫描")
        return []
    
    with open("url.txt", "r", encoding="utf-8") as f:
        lines = f.readlines()
    
    parsed_data = []
    current_id = 1
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # 检查是否为单行URL
        url_match = re.match(r'^(http[s]?://\S+)$', line)
        if url_match:
            url = url_match.group(1)
            # 清除URL末尾的逗号
            url = re.sub(r',$', '', url)  # 新增：移除末尾逗号
            parsed_data.append({
                '序号': current_id,
                'IP地址': '',
                '端口': '',
                '协议': '',
                '状态码': '',
                'URL': url,
                '技术栈': '',
                '页面标题': ''
            })
            current_id += 1
            continue
            
        # 尝试修复可能的格式问题（如URL和协议连在一起）
        line = re.sub(r'(http[s]?://\S+)(TCP/HTTP)', r'\1, \2', line)
            
        # 解析完整格式行：协议, [状态码], [技术栈], URL, [页面标题]
        full_match = re.match(
            r'^(\w+/\w+),\s*\[\s*(\d+)\s*\],\s*\[(.*?)\],\s*(http[s]?://\S+),\s*\[(.*?)\],?$', 
            line
        )
        
        if full_match:
            protocol = full_match.group(1)
            status_code = full_match.group(2)
            technologies = full_match.group(3)
            url = full_match.group(4)
            page_title = full_match.group(5)
            
            # 清除URL末尾的逗号（核心修复）
            url = re.sub(r',$', '', url)  # 新增：移除末尾逗号
            
            # 提取IP和端口
            ip_port_pattern = re.compile(r'http[s]?://([^:/]+):?(\d+)?')
            ip_port_match = ip_port_pattern.search(url)
            ip = ip_port_match.group(1) if ip_port_match else ""
            port = ip_port_match.group(2) if ip_port_match and ip_port_match.group(2) else ""
            
            parsed_data.append({
                '序号': current_id,
                'IP地址': ip,
                '端口': port,
                '协议': protocol,
                '状态码': status_code,
                'URL': url,
                '技术栈': technologies,
                '页面标题': page_title
            })
            current_id += 1
            continue
            
        # 尝试解析其他可能的格式（如缺少页面标题）
        partial_match = re.match(
            r'^(\w+/\w+),\s*\[\s*(\d+)\s*\],\s*\[(.*?)\],\s*(http[s]?://\S+),?\s*$', 
            line
        )
        
        if partial_match:
            protocol = partial_match.group(1)
            status_code = partial_match.group(2)
            technologies = partial_match.group(3)
            url = partial_match.group(4)
            page_title = ""
            
            # 清除URL末尾的逗号（核心修复）
            url = re.sub(r',$', '', url)  # 新增：移除末尾逗号
            
            # 提取IP和端口
            ip_port_pattern = re.compile(r'http[s]?://([^:/]+):?(\d+)?')
            ip_port_match = ip_port_pattern.search(url)
            ip = ip_port_match.group(1) if ip_port_match else ""
            port = ip_port_match.group(2) if ip_port_match and ip_port_match.group(2) else ""
            
            parsed_data.append({
                '序号': current_id,
                'IP地址': ip,
                '端口': port,
                '协议': protocol,
                '状态码': status_code,
                'URL': url,
                '技术栈': technologies,
                '页面标题': page_title
            })
            current_id += 1
            continue
            
        # 如果无法解析，尝试提取URL
        url_in_line = re.search(r'(http[s]?://\S+)', line)
        if url_in_line:
            url = url_in_line.group(1)
            
            # 清除URL末尾的逗号（核心修复）
            url = re.sub(r',$', '', url)  # 新增：移除末尾逗号
            
            # 提取IP和端口
            ip_port_pattern = re.compile(r'http[s]?://([^:/]+):?(\d+)?')
            ip_port_match = ip_port_pattern.search(url)
            ip = ip_port_match.group(1) if ip_port_match else ""
            port = ip_port_match.group(2) if ip_port_match and ip_port_match.group(2) else ""
            
            parsed_data.append({
                '序号': current_id,
                'IP地址': ip,
                '端口': port,
                '协议': '',
                '状态码': '',
                'URL': url,
                '技术栈': '',
                '页面标题': ''
            })
            current_id += 1
            print(f"警告：部分解析行 - {line}")
            continue
            
        # 如果完全无法解析，记录原始行
        print(f"警告：无法解析行 - {line}")
    
    return parsed_data

def generate_excel(data, file_path=None):
    """生成Excel表格并美化"""
    if not data:
        print("没有数据可生成表格")
        return
    
    if file_path is None:
        file_path = f"url_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 保存为Excel文件
    df.to_excel(file_path, index=False, engine='openpyxl')
    
    # 美化Excel表格
    beautify_excel(file_path)
    
    print(f"详细表格已保存到 {file_path}")
    
    # 返回数据用于预览
    return df

def beautify_excel(file_path):
    """美化Excel表格"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 表头样式 - 绿色系
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
        
        # 数据样式
        data_font = Font(color="000000", size=11)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 应用表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 应用数据样式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)  # 限制最大宽度
        
        # 为URL添加超链接样式
        if 'URL' in [cell.value for cell in ws[1]]:
            col_idx = [cell.value for cell in ws[1]].index('URL') + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith(('http', 'https')):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")
        
        # 添加条件格式 - 状态码高亮
        if '状态码' in [cell.value for cell in ws[1]]:
            col_idx = [cell.value for cell in ws[1]].index('状态码') + 1
            status_range = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}"
            
            # 状态码条件格式 - 绿色系渐变
            ws.conditional_formatting.add(status_range,
                ColorScaleRule(start_type='min', start_color='FFD6CC', 
                              mid_type='percentile', mid_value=50, mid_color='FFF6CC',
                              end_type='max', end_color='CCFFCC')
            )
        
        # 设置自动筛选 - 筛选所有列
        if ws.max_row > 1:  # 确保有数据行
            ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
        
        wb.save(file_path)
        print(f"Excel表格美化完成: {file_path}")
    except Exception as e:
        print(f"美化Excel失败: {e}")

def save_urls_to_file(data):
    """将解析的URL保存到url.txt文件中"""
    if not data:
        print("没有URL数据可保存")
        return
    
    try:
        with open("url.txt", "w", encoding="utf-8") as f:
            for item in data:
                url = item.get('URL', '')
                if url:
                    f.write(url + "\n")
        
        print(f"已将 {len(data)} 个URL保存到url.txt")
    except Exception as e:
        print(f"保存URL到文件失败: {e}")

def main():
    """主函数"""
    print("=" * 50)
    print("网络扫描与URL提取工具")
    print("=" * 50)
    
    # 执行扫描
    if run_ts_scan():
        # 解析URL文件
        parsed_data = parse_url_file()
        
        if not parsed_data:
            print("没有解析到任何URL数据！")
            return
        
        # 生成Excel表格（带美化）
        df = generate_excel(parsed_data)
        
        # 保存URL到文件（新增功能）
        save_urls_to_file(parsed_data)
        
        # 打印表格预览
        print("\n表格预览:")
        print(df.head().to_string())
        
        print(f"\n共解析出 {len(parsed_data)} 条记录")
    
    print("\n操作完成!")

if __name__ == "__main__":
    main()