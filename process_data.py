import json
import pandas as pd
import os
import sys
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill, GradientFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 定义期望的列顺序
SPRAY_EXPECTED_COLUMNS = [
    'A', 'B', 'C', 'D', 'E',  # 前5列保持不变（假设E列是URL）
    'F', 'G', 'H', 'I', 'J',  # J列是状态码
    'K', 'L', 'M', 'N', 'O',  # O列可能是directurl
    # 添加其他需要的列...
]

# 定义列标题映射（如果需要重命名列）
COLUMN_MAPPING = {
    'directurl': 'Direct URL',  # 示例：将directurl列重命名为Direct URL
    # 添加其他需要的映射...
}

def extract_names(data):
    """从嵌套字典中提取所有'name'字段并拼接"""
    names = []
    try:
        if isinstance(data, str):
            data = json.loads(data)
        for key, value in data.items():
            if isinstance(value, dict) and 'name' in value:
                names.append(value['name'])
    except (json.JSONDecodeError, AttributeError, TypeError):
        return ""
    return ' | '.join(names)

def filter_valid_urls(df, status_code_col='J', url_col='E'):
    """筛选状态码为200的URL记录"""
    if status_code_col not in df.columns or url_col not in df.columns:
        print(f"警告: 未找到列 {status_code_col} 或 {url_col}，跳过筛选")
        return df
    
    return df[(df[status_code_col] == 200) & (df[url_col].notna())]

def beautify_spray_excel(file_path):
    """美化spray生成的Excel表格"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        # 数据样式
        data_font = Font(color="000000", size=11)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 应用表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 应用数据样式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)  # 限制最大宽度
        
        # 添加条件格式 - 状态码高亮
        if 'J' in [get_column_letter(cell.column) for cell in ws[1]]:
            col_idx = [get_column_letter(cell.column) for cell in ws[1]].index('J') + 1
            status_range = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}"
            
            # 状态码条件格式
            ws.conditional_formatting.add(status_range,
                ColorScaleRule(start_type='min', start_color='FFC7CE', 
                              mid_type='percentile', mid_value=50, mid_color='FFFFCC',
                              end_type='max', end_color='C6EFCE')
            )
        
        wb.save(file_path)
        print(f"Spray Excel表格美化完成: {file_path}")
    except Exception as e:
        print(f"美化Spray Excel失败: {e}")

def beautify_ehole_excel(file_path):
    """深度美化ehole生成的Excel表格"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 检查工作表是否为空
        if ws.max_row <= 1:
            print(f"警告: ehole结果表格为空: {file_path}")
            return
        
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        
        # 数据样式
        data_font = Font(color="000000", size=12)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 应用表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 应用数据样式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 60)  # 限制最大宽度
        
        # 为URL添加超链接样式
        url_cols = ['URL', 'url', '网址']  # 可能的URL列名
        for col_name in url_cols:
            if col_name in [cell.value for cell in ws[1]]:
                col_idx = [cell.value for cell in ws[1]].index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith(('http', 'https')):
                        cell.hyperlink = cell.value
                        cell.font = Font(color="0563C1", underline="single")
        
        # 添加条件格式 - 风险级别高亮
        risk_cols = ['Risk', '风险等级', '危险程度']  # 可能的风险列名
        for col_name in risk_cols:
            if col_name in [cell.value for cell in ws[1]]:
                col_idx = [cell.value for cell in ws[1]].index(col_name) + 1
                risk_range = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}"
                
                # 高风险(红色)
                ws.conditional_formatting.add(risk_range,
                    CellIsRule(operator='containsText', formula=['"高"'], 
                              fill=PatternFill(bgColor='FFC7CE'),
                              font=Font(color='9C0006'))
                )
                
                # 中风险(黄色)
                ws.conditional_formatting.add(risk_range,
                    CellIsRule(operator='containsText', formula=['"中"'], 
                              fill=PatternFill(bgColor='FFEB9C'),
                              font=Font(color='9C5700'))
                )
                
                # 低风险(绿色)
                ws.conditional_formatting.add(risk_range,
                    CellIsRule(operator='containsText', formula=['"低"'], 
                              fill=PatternFill(bgColor='C6EFCE'),
                              font=Font(color='006100'))
                )
                break
        
        # 创建数据透视表工作表
        if ws.max_row > 10:  # 数据足够多时才创建
            try:
                # 创建数据透视表
                from openpyxl.utils import get_column_letter
                from openpyxl.worksheet.datavalidation import DataValidation
                
                # 获取数据范围
                data_range = f"'{ws.title}'!$A$1:${get_column_letter(ws.max_column)}${ws.max_row}"
                
                # 创建数据透视表工作表
                pivot_ws = wb.create_sheet(title="数据透视表")
                
                # 添加数据透视表标题
                pivot_ws['A1'] = "指纹识别结果统计"
                pivot_ws['A1'].font = Font(bold=True, size=16)
                
                # 准备数据透视表
                from openpyxl.pivot.table import PivotTable, PivotField
                from openpyxl.pivot.fields import PageField
                
                # 创建数据透视表对象
                pt = PivotTable(
                    srcRange=data_range,
                    dest=f"'{pivot_ws.title}'!$A$3",
                    name="指纹统计"
                )
                
                # 添加行字段（假设第一列是软件/系统名称）
                pt.addRow('A')
                
                # 添加列字段（假设第二列是版本）
                if ws.max_column >= 2:
                    pt.addColumn('B')
                
                # 添加数据字段（计数）
                pt.addData('A', function='count')
                
                # 添加数据透视表到工作表
                pivot_ws.add_pivot(pt)
                
                # 美化数据透视表
                for row in pivot_ws.iter_rows(min_row=3, max_row=3):
                    for cell in row:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")
                
                # 添加筛选下拉框
                if ws.max_column >= 3:
                    # 获取第三列的唯一值作为筛选条件
                    filter_values = list({ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)})
                    
                    # 创建数据验证对象
                    dv = DataValidation(type="list", formula1='"{}"'.format(','.join([str(v) for v in filter_values if v])))
                    
                    # 添加到单元格
                    dv.add(pivot_ws['D1'])
                    pivot_ws.add_data_validation(dv)
                    
                    # 添加筛选标签
                    pivot_ws['C1'] = "筛选:"
                    pivot_ws['C1'].font = Font(bold=True)
                
                print(f"已为ehole结果添加数据透视表")
            except Exception as e:
                print(f"创建数据透视表失败: {e}")
        
        # 添加汇总信息
        summary_ws = wb.create_sheet(title="汇总信息")
        summary_ws['A1'] = "指纹识别结果汇总"
        summary_ws['A1'].font = Font(bold=True, size=16)
        
        summary_ws['A3'] = "总记录数:"
        summary_ws['B3'] = ws.max_row - 1  # 减去表头
        
        # 保存文件
        wb.save(file_path)
        print(f"Ehole Excel表格深度美化完成: {file_path}")
    except Exception as e:
        print(f"美化Ehole Excel失败: {e}")

def process_data(input_file, output_file):
    """处理JSON输入文件，生成Excel和TXT输出"""
    try:
        # 检查输入文件类型
        file_ext = os.path.splitext(input_file)[1].lower()
        
        if file_ext == '.json':
            # 处理spray生成的JSON文件
            print(f"开始处理spray结果: {input_file}")
            
            # 读取JSON数据
            data_list = []
            with open(input_file, 'r', encoding='utf-8') as f:
                for line in f:
                    try:
                        data = json.loads(line.strip())
                        data_list.append(data)
                    except json.JSONDecodeError:
                        print(f"警告: 无法解析JSON行: {line[:50]}...")
            
            if not data_list:
                print(f"错误: 文件 {input_file} 中没有有效JSON数据")
                return
            
            # 转换为DataFrame
            df = pd.DataFrame(data_list)
            
            # 重命名列（如果需要）
            df = df.rename(columns=COLUMN_MAPPING)
            
            # 检测并删除redirect_url列
            if 'redirect_url' in df.columns:
                print(f"检测到'redirect_url'列，已删除")
                df = df.drop(columns=['redirect_url'])
            else:
                print(f"未检测到'redirect_url'列")
            
            # 处理O列数据（提取name字段）
            if 'O' in df.columns:
                df['O'] = df['O'].apply(extract_names)
            
            # 调整列顺序
            # 获取实际存在的列
            existing_columns = df.columns.tolist()
            
            # 构建期望的列顺序，对于不存在的列使用None占位
            ordered_columns = []
            for col_letter in SPRAY_EXPECTED_COLUMNS:
                # 将列字母转换为索引（A→0, B→1, ...）
                col_index = ord(col_letter.upper()) - 65
                if col_index < len(existing_columns):
                    ordered_columns.append(existing_columns[col_index])
                else:
                    # 如果索引超出范围，添加None
                    ordered_columns.append(None)
            
            # 过滤掉None，只保留实际存在的列
            ordered_columns = [col for col in ordered_columns if col is not None]
            
            # 确保所有列都被包含
            for col in existing_columns:
                if col not in ordered_columns:
                    ordered_columns.append(col)
            
            # 重新排列DataFrame的列
            df = df[ordered_columns]
            
            # 筛选有效URL（状态码200）
            if 'J' in df.columns and 'E' in df.columns:
                valid_df = filter_valid_urls(df)
                print(f"筛选后保留 {len(valid_df)}/{len(df)} 条记录")
            else:
                valid_df = df
                print(f"未找到状态码列，保留所有 {len(df)} 条记录")
            
            # 保存Excel文件
            valid_df.to_excel(output_file, index=False)
            print(f"Excel文件已保存: {output_file}")
            
            # 提取第五列（索引4）的URL并保存为TXT
            if len(valid_df.columns) >= 5:
                url_column = valid_df.iloc[:, 4]
                urls = url_column.dropna().tolist()
                
                if urls:
                    txt_output = os.path.splitext(output_file)[0] + ".txt"
                    with open(txt_output, 'w', encoding='utf-8') as f:
                        f.write('\n'.join(urls))
                    print(f"已提取 {len(urls)} 个URL保存到: {txt_output}")
                else:
                    print("警告: 未找到有效URL")
            else:
                print(f"错误: 数据不足5列，无法提取URL")
            
            # 美化spray生成的Excel
            beautify_spray_excel(output_file)
        
        elif file_ext in ['.xlsx', '.xls']:
            # 处理ehole生成的Excel文件
            print(f"开始美化ehole结果: {input_file}")
            
            # 直接美化Excel文件
            if input_file != output_file:
                # 复制文件
                shutil.copy2(input_file, output_file)
            
            # 深度美化ehole结果（关键修复点：恢复了对beautify_ehole_excel的调用）
            beautify_ehole_excel(output_file)
        
        else:
            print(f"错误: 不支持的文件类型: {file_ext}")
            
    except Exception as e:
        print(f"处理文件时出错: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("用法: python process_data.py <输入JSON/Excel文件> <输出Excel文件>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    process_data(input_file, output_file)