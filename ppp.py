import os
import re
import sys
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Border, Side, Alignment, Font, PatternFill, NamedStyle
)
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

# 定义全局样式
HEADER_STYLE = {
    "font": Font(bold=True, color="FFFFFF", size=12),
    "fill": PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid"),
    "alignment": Alignment(horizontal='center', vertical='center')
}

DATA_STYLE = {
    "font": Font(color="000000", size=11),
    "border": Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin')),
    "alignment": Alignment(horizontal='left', vertical='center', wrap_text=True)
}

CONDITIONAL_FORMATTING = {
    "status_code": {
        "200": "00FF00",
        "404": "FFFF00",
        "500": "FF0000"
    },
    "source_colors": {
        "状态": "F0F0F0",
        "指纹": "E0FFFF",
        "URL": "FFF0F5"
    }
}


def parse_portscan_file(file_path="port.txt"):
    """解析port.txt文件，支持多种格式"""
    if not os.path.exists(file_path):
        print(f"错误: 文件 '{file_path}' 不存在")
        return []

    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    parsed_data = []
    current_id = 1
    host_pattern = r'([^:\s,\[\]]+)'
    patterns = {
        "status": re.compile(rf'^{host_pattern}:(\d+)\s+(\w+)$'),
        "fingerprint": re.compile(
            rf'^([A-Z/]+),\s*,\s*\[(.*?)\],\s*{host_pattern}:(\d+),\s*\[(.*?)\],?$'
        ),
        "url": re.compile(
            r'^([A-Z/]+),\s*\[(\d+)\],\s*\[(.*?)\],\s*(http[s]?://\S+),\s*\[(.*?)\],?$'
        ),
        "empty_fingerprint": re.compile(
            rf'^([A-Z/]+),\s*,\s*,\s*{host_pattern}:(\d+),\s*\[.*\],?$'
        )
    }

    for line in lines:
        line = line.strip().rstrip(',')
        if not line:
            continue

        # 按优先级匹配不同格式
        match = patterns["status"].match(line)
        if match:
            _parse_status_line(match, parsed_data, current_id)
            current_id += 1
            continue

        match = patterns["fingerprint"].match(line)
        if match:
            _parse_fingerprint_line(match, parsed_data, current_id)
            current_id += 1
            continue

        match = patterns["empty_fingerprint"].match(line)  # 处理空指纹的端口行
        if match:
            protocol = match.group(1)
            ip = match.group(2)
            port = match.group(3)
            parsed_data.append({
                "序号": current_id,
                "IP地址": ip,
                "端口": port,
                "协议": protocol,
                "服务信息": "open",
                "指纹信息": "",
                "URL": "",
                "来源": "指纹"
            })
            current_id += 1
            continue

        match = patterns["url"].match(line)
        if match:
            _parse_url_line(match, parsed_data, current_id)
            current_id += 1
            continue

        print(f"警告：无法解析行 - {line}")

    return parsed_data


def _parse_status_line(match, parsed_data, current_id):
    """解析纯端口状态行"""
    ip, port, status = match.groups()
    parsed_data.append({
        "序号": current_id,
        "IP地址": ip,
        "端口": port,
        "服务信息": status,
        "来源": "状态"
    })


def _parse_fingerprint_line(match, parsed_data, current_id):
    """解析端口指纹行"""
    protocol, component, ip, port, version = match.groups()
    fingerprint = f"{component} ({version})" if version else component
    parsed_data.append({
        "序号": current_id,
        "IP地址": ip,
        "端口": port,
        "协议": protocol,
        "服务信息": "open",
        "指纹信息": fingerprint,
        "来源": "指纹"
    })


def _parse_url_line(match, parsed_data, current_id):
    """解析URL行"""
    protocol, status_code, fingerprint, url, title = match.groups()
    ip_port = re.search(r'http[s]?://([^:/]+):?(\d+)?', url)
    ip = ip_port.group(1) if ip_port else ""
    port = ip_port.group(2) if ip_port and ip_port.group(2) else ""
    parsed_data.append({
        "序号": current_id,
        "IP地址": ip,
        "端口": port,
        "协议": protocol,
        "服务信息": "HTTP服务",
        "指纹信息": fingerprint,
        "URL": url,
        "状态码": status_code,
        "页面标题": title,
        "来源": "URL"
    })


def generate_excel(data, file_path=None):
    """生成并美化Excel表格"""
    if not data:
        print("没有数据可生成表格")
        return

    columns = [
        "序号", "IP地址", "端口", "协议", "服务信息", "指纹信息",
        "URL", "状态码", "页面标题", "来源"
    ]
    df = pd.DataFrame(data, columns=columns).fillna("")

    if file_path is None:
        file_path = f"port_scan_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, index=False, sheet_name="扫描结果")
        ws = writer.sheets["扫描结果"]
        _apply_styles(ws)

    print(f"表格已保存至: {file_path}")


def _apply_styles(ws):
    """应用全局样式"""
    # 应用表头样式
    for cell in ws[1]:
        cell.font = HEADER_STYLE["font"]
        cell.fill = HEADER_STYLE["fill"]
        cell.alignment = HEADER_STYLE["alignment"]

    # 应用数据样式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = DATA_STYLE["font"]
            cell.border = DATA_STYLE["border"]
            cell.alignment = DATA_STYLE["alignment"]

    # 自动调整列宽
    for col in ws.columns:
        max_width = max(len(str(cell.value)) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_width, 30)

    # 状态码条件格式
    if "状态码" in [cell.value for cell in ws[1]]:
        col_idx = [cell.value for cell in ws[1]].index("状态码") + 1
        for code, color in CONDITIONAL_FORMATTING["status_code"].items():
            ws.conditional_formatting.add(
                f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}",
                CellIsRule(operator="equal", formula=[code], fill=PatternFill(start_color=color, fill_type="solid"))
            )

    # 来源列背景色
    if "来源" in [cell.value for cell in ws[1]]:
        col_idx = [cell.value for cell in ws[1]].index("来源") + 1
        for source, color in CONDITIONAL_FORMATTING["source_colors"].items():
            ws.conditional_formatting.add(
                f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}",
                CellIsRule(operator="equal", formula=[f'"{source}"'], fill=PatternFill(start_color=color, fill_type="solid"))
            )

    # 添加冻结窗格
    ws.freeze_panes = "A2"


def main():
    print("=" * 60)
    print("端口扫描报告生成工具 (增强美观版)")
    print("支持解析纯端口状态、端口指纹、URL格式，并自动美化表格")
    print("=" * 60)

    data = parse_portscan_file()

    if not data:
        print("解析失败或无有效数据！")
        return

    generate_excel(data)

    print("\n操作完成！结果文件已生成")


if __name__ == "__main__":
    main()